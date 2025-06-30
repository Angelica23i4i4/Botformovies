from aiogram import Bot, types, executor, Dispatcher
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.dispatcher.filters.state import StatesGroup, State
from openpyxl import load_workbook


class ClientState(StatesGroup):
    START_REQUEST = State()
    GENRE_SELECTED = State()
    GENRE_SELECTED2 = State()
    DESCRIPTION = State()
    EXACTING = State()


TOKEN = "The bot token"


# Redisstorage
# MongoStorage
storage = MemoryStorage()
bot = Bot(TOKEN)
db = Dispatcher(bot, storage=storage)

# Хранит на каком этапе диалога у нас клиент


@db.message_handler(commands=['start'])
async def command_start_handler(message: types.Message, state: FSMContext) -> None:
    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton('Приключенческие'))
    markup.add(KeyboardButton('Боевики'))
    markup.add(KeyboardButton('Детектив'))
    markup.add(KeyboardButton('Биографические'))
    markup.add(KeyboardButton('Исторические'))
    markup.add(KeyboardButton('Триллеры'))
    markup.add(KeyboardButton('Комедийные'))
    markup.add(KeyboardButton('Драмы'))
    markup.add(KeyboardButton('Фэнтези'))
    markup.add(KeyboardButton('Фантастические'))
    await message.answer(f"Привет! Какие сериалы хотела бы посмотреть?", reply_markup=markup)
    await state.set_state(ClientState.START_REQUEST)


@db.message_handler(state=ClientState.START_REQUEST)
async def ask(message: types.Message, state: FSMContext) -> None:
    user = message.text
    if user == 'Нет':
        await state.update_data(GENRE3='')
    else:
        await state.update_data(GENRE1=user)

    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton('Приключенческие'))
    markup.add(KeyboardButton('Боевики'))
    markup.add(KeyboardButton('Детектив'))
    markup.add(KeyboardButton('Биографические'))
    markup.add(KeyboardButton('Исторические'))
    markup.add(KeyboardButton('Триллеры'))
    markup.add(KeyboardButton('Комедийные'))
    markup.add(KeyboardButton('Драмы'))
    markup.add(KeyboardButton('Фантастические'))
    markup.add(KeyboardButton('Фэнтези'))
    markup.add(KeyboardButton('Нет'))
    await bot.send_message(chat_id=message.from_id, text='Хочешь добавить еще жанр?', reply_markup=markup)
    await state.set_state(ClientState.GENRE_SELECTED)


@db.message_handler(state=ClientState.GENRE_SELECTED)
async def ask2(message: types.Message, state: FSMContext) -> None:
    user = message.text
    if user == 'Нет':
        await state.update_data(GENRE3='')
    else:
        await state.update_data(GENRE2=user)
    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton('Приключенческие'))
    markup.add(KeyboardButton('Боевики'))
    markup.add(KeyboardButton('Детектив'))
    markup.add(KeyboardButton('Биографические'))
    markup.add(KeyboardButton('Исторические'))
    markup.add(KeyboardButton('Триллеры'))
    markup.add(KeyboardButton('Комедийные'))
    markup.add(KeyboardButton('Драмы'))
    markup.add(KeyboardButton('Фантастические'))
    markup.add(KeyboardButton('Фэнтези'))
    markup.add(KeyboardButton('Нет'))
    await bot.send_message(chat_id=message.chat.id, text='Последняя попытка выбрать еще жанр')
    await state.set_state(ClientState.GENRE_SELECTED2)


@db.message_handler(state=ClientState.GENRE_SELECTED2)
async def ask2(message: types.Message, state: FSMContext) -> None:
    user = message.text
    if user == 'Нет':
        await state.update_data(GENRE3='')
    else:
        await state.update_data(GENRE3=user)
    await bot.send_message(chat_id=message.chat.id, text='напиши ключевые слова для описания cериала')
    await state.set_state(ClientState.DESCRIPTION)


@db.message_handler(state=ClientState.DESCRIPTION)
async def ask3(message: types.Message, state: FSMContext) -> None:
    user = message.text
    list_user = user.split(' ')
    user_get_data = await state.get_data()
    data_des = []
    sent = []
    for genre in user_get_data.values():
        if genre != '':
            print(user_get_data.values(), list_user)
            n = f'{genre}.xlsx'
            book = load_workbook(filename=n)
            sheet = book['Sheet1']
            for item in range(1, 60):
                list1 = sheet['B' + str(item)].value.replace('\n', '').split('  ')
                list2 = [genre, '']
                if user_get_data.values() == list1:
                    # await bot.send_message(chat_id=message.chat.id, text=sheet['E' + str(item)].value)
                    data_des.append(sheet['E' + str(item)].value)
                elif (user_get_data['GENRE1'] not in list2) and (user_get_data['GENRE1'] in list1):
                    # await bot.send_message(chat_id=message.chat.id, text=sheet['E' + str(item)].value)
                    data_des.append(sheet['E' + str(item)].value)
                elif (user_get_data['GENRE2'] not in list2) and (user_get_data['GENRE2'] in list1):
                    # await bot.send_message(chat_id=message.chat.id, text=sheet['E' + str(item)].value)
                    data_des.append(sheet['E' + str(item)].value)
                elif (user_get_data['GENRE3'] not in list2) and (user_get_data['GENRE3'] in list1):
                    # await bot.send_message(chat_id=message.chat.id, text=sheet['E' + str(item)].value)
                    data_des.append(sheet['E' + str(item)].value)
                else:
                    continue

            if 'Нет' not in list_user:
                for link in data_des:
                    plus = []
                    for word in list_user:
                        n = f'{genre}.xlsx'
                        book = load_workbook(filename=n)
                        sheet = book['Sheet1']
                        for item in range(1, 60):
                            des = sheet['D' + str(item)].value.replace('Описание:', '')
                            if sheet['E' + str(item)].value == link and word in des and len(word) > 3:
                                plus.append(word)
                            else:
                                continue
                if len(plus) >= 2 and link not in sent:
                    sent.append(link)
                    await bot.send_message(chat_id=message.chat.id, text=link)
                else:
                    await bot.send_message(chat_id=message.chat.id, text='По вашему запросу ничего не найдено')
        else:
            continue


if __name__ == "__main__":
    executor.start_polling(db)
